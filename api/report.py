from http.server import BaseHTTPRequestHandler
import json
import io

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        # Enable CORS for frontend integration
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        try:
            content_length = int(self.headers.get('content-length', 0))
            post_data = self.rfile.read(content_length)
            
            try:
                data = json.loads(post_data.decode('utf-8'))
            except Exception:
                data = {}

            report_format = data.get('format', 'pdf').lower()
            rows = data.get('rows', [])
            title = data.get('title', 'EcoSphere ESG Report')
            
            if not rows:
                # Return empty/default rows if none provided
                rows = [{"Status": "No data available for export"}]

            # Get columns from the first row keys
            columns = list(rows[0].keys())

            if report_format == 'xlsx':
                buf = self.generate_excel(title, columns, rows)
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                filename = f"{title.replace(' ', '_').lower()}.xlsx"
            else:
                buf = self.generate_pdf(title, columns, rows)
                content_type = 'application/pdf'
                filename = f"{title.replace(' ', '_').lower()}.pdf"

            self.send_response(200)
            self.send_header('Content-Type', content_type)
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(buf.read())

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            err_response = {"error": str(e)}
            self.wfile.write(json.dumps(err_response).encode('utf-8'))

    def generate_pdf(self, title, columns, rows):
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors

        buf = io.BytesIO()
        # Set up document with 0.5 inch margins
        doc = SimpleDocTemplate(
            buf, 
            pagesize=A4, 
            leftMargin=36, 
            rightMargin=36, 
            topMargin=36, 
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        # Define clean, curated color palette
        primary_color = colors.HexColor("#0D9488")   # Teal 600
        secondary_color = colors.HexColor("#0F172A") # Slate 900
        light_bg = colors.HexColor("#F8FAFC")        # Slate 50
        border_color = colors.HexColor("#E2E8F0")    # Slate 200

        # Custom paragraph styles
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            textColor=primary_color,
            spaceAfter=15
        )
        
        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.white
        )
        
        table_cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=secondary_color
        )

        elements = []
        
        # Add Title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 10))

        # Build Table Data
        table_data = []
        # Header Row
        header_row = [Paragraph(col, table_header_style) for col in columns]
        table_data.append(header_row)

        # Value Rows
        for row in rows:
            val_row = []
            for col in columns:
                val = str(row.get(col, ''))
                val_row.append(Paragraph(val, table_cell_style))
            table_data.append(val_row)

        # Style the Table
        # Width distribution: automatically divide remaining width among columns
        page_width = A4[0] - 72 # total A4 width - margins (36 left + 36 right)
        col_width = page_width / len(columns)
        
        t = Table(table_data, colWidths=[col_width] * len(columns))
        
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_color),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
        ])
        
        # Add alternating row colors
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                t_style.add('BACKGROUND', (0, i), (-1, i), light_bg)

        t.setStyle(t_style)
        elements.append(t)
        
        doc.build(elements)
        buf.seek(0)
        return buf

    def generate_excel(self, title, columns, rows):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ESG Report"
        
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        title_font = Font(name="Calibri", size=16, bold=True, color="0D9488")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
        cell_font = Font(name="Calibri", size=11)
        
        thin_side = Side(border_style="thin", color="E2E8F0")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Write Title
        ws.append([title])
        ws.cell(row=1, column=1).font = title_font
        ws.row_dimensions[1].height = 30
        ws.append([]) # Empty row

        # Write Headers
        ws.append(columns)
        ws.row_dimensions[3].height = 24
        
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

        # Write Data
        start_row = 4
        for row_idx, row in enumerate(rows, start=start_row):
            row_data = [row.get(col, '') for col in columns]
            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 20
            
            # Formatting cells
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Alternate row coloring
                if (row_idx - start_row) % 2 == 1:
                    cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                # Avoid taking title row for width calculation
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(buf)
        buf.seek(0)
        return buf

if __name__ == '__main__':
    from http.server import HTTPServer
    server = HTTPServer(('localhost', 3000), handler)
    print("Local report API server running at http://localhost:3000/api/report")
    print("Press Ctrl+C to stop.")
    server.serve_forever()
